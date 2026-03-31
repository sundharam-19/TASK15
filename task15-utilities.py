import openpyxl
import datetime

class ExcelUtils:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook[sheet_name]

    def get_test_data(self):
        data = []
        # Iterate through rows starting from the second row (skipping header)
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            # Map values to dictionary using header names
            row_data = {
                "Username": row[0],
                "Password": row[1],
                "Date": row[2],
                "Time of Test": row[3],
                "Name of Tester": row[4],
                "Test Result": row[5]
            }
            data.append(row_data)
        return data

    def write_test_result(self, row_index, result, tester_name):
        # Update Date, Time, Tester Name, and Test Result in the correct columns
        current_date = datetime.date.today().strftime("%Y-%m-%d")
        current_time = datetime.datetime.now().strftime("%H:%M:%S")

        # Column indices (assuming A=1, B=2, ... F=6)
        self.sheet.cell(row=row_index + 2, column=3, value=current_date)
        self.sheet.cell(row=row_index + 2, column=4, value=current_time)
        self.sheet.cell(row=row_index + 2, column=5, value=tester_name)
        self.sheet.cell(row=row_index + 2, column=6, value=result)
        self.workbook.save(self.file_path)